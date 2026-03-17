#!/usr/bin/env python3
"""Write BMS extraction results to an Excel workbook.

Usage:
    python write_xlsx.py <output_dir>

Reads:
    <output_dir>/<site_name>_sitemodel.json  — extracted site model
    references/equipment-types.md             — master equipment type list (relative to script dir)

Writes:
    <output_dir>/<site_name>_assetregister.xlsx  — workbook with 3 tabs
"""

import json
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl


def parse_equipment_types(md_path: str) -> list[str]:
    """Extract equipment type names from the markdown reference file."""
    types = []
    with open(md_path) as f:
        for line in f:
            m = re.match(r"^\d+\.\s+(.+)$", line.strip())
            if m:
                types.append(m.group(1))
    return types


def write_xlsx(output_dir: str) -> None:
    # Find the *_sitemodel.json file in the output directory
    sitemodel_files = [f for f in os.listdir(output_dir) if f.endswith("_sitemodel.json")]
    if not sitemodel_files:
        print(f"Error: no *_sitemodel.json found in {output_dir}")
        sys.exit(1)
    model_path = os.path.join(output_dir, sitemodel_files[0])
    with open(model_path) as f:
        model = json.load(f)

    script_dir = os.path.dirname(os.path.abspath(__file__))
    eq_types_path = os.path.join(script_dir, "equipment-types.md")
    equipment_types = parse_equipment_types(eq_types_path)

    site_name = model["site"].lower().replace(" ", "_")
    wb = openpyxl.Workbook()

    # Tab 1: levels_and_zones
    ws1 = wb.active
    ws1.title = "levels_and_zones"
    ws1.append(["level", "equipments", "zones", "source_url"])
    for level in model.get("levels", []):
        equipments = level.get("equipments", "")
        if isinstance(equipments, list):
            equipments = ", ".join(equipments)
        zones = level.get("zones", "")
        if isinstance(zones, list):
            zones = ", ".join(zones)
        ws1.append([level["name"], equipments, zones, level.get("source_url", "")])

    # Tab 2: equipment_list
    ws2 = wb.create_sheet("equipment_list")
    ws2.append(["equipment_name", "level_select", "zone_select", "equipment_type_select", "source_url"])
    for eq in model.get("equipment_list", []):
        ws2.append([
            eq["equipment_name"],
            eq["level_select"],
            eq["zone_select"],
            eq["equipment_type_select"],
            eq.get("source_url", ""),
        ])

    # Tab 3: equipment_types
    ws3 = wb.create_sheet("equipment_types")
    ws3.append(["equipment_types"])
    for et in equipment_types:
        ws3.append([et])

    out_path = os.path.join(output_dir, f"{site_name}_assetregister.xlsx")
    wb.save(out_path)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2 or sys.argv[1] in ("-h", "--help"):
        print(__doc__.strip())
        sys.exit(0)
    write_xlsx(sys.argv[1])
